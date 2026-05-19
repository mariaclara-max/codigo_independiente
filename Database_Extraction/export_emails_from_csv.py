import os
import re
import csv
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

# Configuración de rutas
SOURCE_DIR = Path("Source")
RESULTS_DIR = Path("Results")

# Crear directorio Results si no existe
RESULTS_DIR.mkdir(exist_ok=True)

# Patrón para validar y extraer emails
EMAIL_PATTERN = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

def extract_emails_from_cell(cell_value):
    """Extrae emails de una celda individual."""
    if cell_value is None:
        return []
    
    cell_str = str(cell_value).strip()
    if not cell_str or cell_str.lower() == 'none':
        return []
    
    # Buscar todos los emails en la celda
    emails = re.findall(EMAIL_PATTERN, cell_str)
    return emails

def process_xlsx_file(file_path):
    """Procesa un archivo XLSX y extrae todos los emails."""
    emails = set()
    
    try:
        # Usar openpyxl directamente
        workbook = load_workbook(file_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Iterar sobre todas las celdas
            for row in sheet.iter_rows(values_only=True):
                for cell_value in row:
                    found_emails = extract_emails_from_cell(cell_value)
                    emails.update(found_emails)
        
        workbook.close()
    
    except Exception as e:
        print(f"Error procesando {file_path}: {e}")
    
    return emails

def get_domain_from_email(email):
    """Extrae el dominio de una dirección email."""
    return email.split('@')[1].lower()

def main():
    print("Iniciando extracción de emails...")
    
    all_emails = set()
    
    # Buscar todos los archivos .xlsx en Source y sus subcarpetas
    xlsx_files = sorted(SOURCE_DIR.rglob("*.xlsx"))
    print(f"Encontrados {len(xlsx_files)} archivos xlsx\n")
    
    for idx, xlsx_file in enumerate(xlsx_files, 1):
        print(f"[{idx}/{len(xlsx_files)}] Procesando: {xlsx_file}")
        emails = process_xlsx_file(xlsx_file)
        all_emails.update(emails)
        print(f"  → {len(emails)} emails únicos en este archivo")
    
    print(f"\n{'='*50}")
    print(f"Total de emails únicos encontrados: {len(all_emails)}")
    print(f"{'='*50}\n")
    
    if not all_emails:
        print("No se encontraron emails. Verifica los archivos source.")
        return
    
    # Guardar emails únicos en CSV
    emails_list = sorted(list(all_emails))
    output_emails = RESULTS_DIR / "emails_list.csv"
    with open(output_emails, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Email'])
        for email in emails_list:
            writer.writerow([email])
    print(f"✓ Guardado: {output_emails}")
    
    # Contar emails por dominio
    domains = [get_domain_from_email(email) for email in all_emails]
    domain_counts = Counter(domains)
    
    # Guardar conteo por dominio en CSV
    output_domains = RESULTS_DIR / "emails_per_domain.csv"
    with open(output_domains, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Dominio', 'Cantidad'])
        for domain, count in domain_counts.most_common():
            writer.writerow([domain, count])
    print(f"✓ Guardado: {output_domains}")
    
    # Estadísticas finales
    print(f"\n=== Resumen ===")
    print(f"Total de emails únicos: {len(all_emails)}")
    print(f"Total de dominios únicos: {len(domain_counts)}")

if __name__ == "__main__":
    main()
